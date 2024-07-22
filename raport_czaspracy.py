import sys
import psycopg2.extras
import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Alignment, Side, Font, Color,colors
from datetime import datetime, timedelta


DB_HOST = "*"
DB_NAME = "*"
DB_USER = "*"
DB_PASS = "*"


stockdate = sys.argv[1]
#stockdate = "2023-10-10"


conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

DEC2FLOAT = psycopg2.extensions.new_type(psycopg2.extensions.DECIMAL.values, 'DEC2FLOAT',
                                         lambda value, curs: float(value) if value is not None else None)
psycopg2.extensions.register_type(DEC2FLOAT)
headers = ['Data', 'Imię i Nazwisko', 'Login', 'BIBI', 'Departament', 'Funkcja', 'HC/AGENCY', 'Pozycja', 'Zmiana',
           'StartPracy', 'Koniec Pracy', 'Czas Pracy', 'Przerwa']

# Define headers for "Pracownicy" sheet
pracownicy_headers = ['Login','Data','Imię i Nazwisko', 'BIBI', 'Departament', 'Funkcja', 'HC/AGENCY', 'Pozycja', 'Zmiana',
           'StartPracy', 'Koniec Pracy', 'Czas Pracy', 'Przerwa śniadaniowa','Komentarz','Rozpoczęcie','Zakończenie','Sprawdzenie czasu pracy','Sprawdzanie śniadaniowa','Podsumowanie ']


wb = openpyxl.Workbook()

page = wb.create_sheet("Wsad")
page.append(headers)  # Write the headers to the first line


cur.execute(
    """
    SELECT
        pa_d.dzien_pracy AS Data,
        wykpa_d.p_imie || ' ' || wykpa_d.p_nazwisko AS IiN,
        wykpa_d.p_login AS login,
        m11wykpa_d.p_value1527 AS BIBI,
        m11wykpa_d.p_value1447 AS department,
        m11wykpa_d.p_value1449 AS funkcja,
        m11wykpa_d.p_value1448 AS hcorag,
        m11wykpa_d.p_value1479 AS pozycja,
        wykpa_d.zm_idzmiany,
        TO_CHAR(pa_d.data_start, 'HH24:MI'),
        TO_CHAR(pa_d.data_stop, 'HH24:MI'),
        EXTRACT(HOUR FROM (pa_d.data_stop - pa_d.data_start)) || ':' || EXTRACT(MINUTE FROM (pa_d.data_stop - pa_d.data_start)) AS roznica_godzina_minuty,
        pa_d.rcp_akcja AS akcja
    FROM
        (
            SELECT
                ((tpra.p_idpracownika::TEXT) || (tpra.pra_datastart::date)::TEXT) AS search_pole,
                tpra.p_idpracownika,
                SUM(tpra.pra_rbh) AS suma_rbh,
                SUM(tpra.pra_rbh * (tpra.pra_zaangazpracownika / 100)) AS suma_rbh_ef,
                MIN(tpra.pra_datastart) AS data_start,
                MAX(tpra.pra_datastop) AS data_stop,
                tpra.pra_datastart::DATE AS dzien_pracy,
                NULL::INT AS rcp_akcja,
                0 AS zrodlo
            FROM
                tg_praceall AS tpra
            WHERE
                (tpra.pra_flaga & 2 = 2)
                AND (tpra.pra_datastart::DATE <= %s)
                AND (tpra.pra_datastart::DATE >= %s)
            GROUP BY
                tpra.p_idpracownika,
                dzien_pracy
            UNION
            SELECT
                ((rcp.p_idpracownika::TEXT) || (rcp.rcp_czaswydarzenia::DATE)::TEXT || (rcp_idwydarzenia::TEXT)) AS search_pole,
                rcp.p_idpracownika,
                (CASE WHEN rcp.rcp_typwydarzenia IN (1, 3) THEN (EXTRACT(EPOCH FROM COALESCE(rcp.czaswydarzenia_next, now()) - rcp.rcp_czaswydarzenia) / 3600) ELSE 0 END) AS suma_rbh,
                0 AS suma_rbh_ef,
                rcp.rcp_czaswydarzenia AS data_start,
                COALESCE(rcp.czaswydarzenia_next, now()) AS data_stop,
                rcp.rcp_czaswydarzenia::DATE AS dzien_pracy,
                rcp.rcp_typwydarzenia AS rcp_akcja,
                1 AS zrodlo
            FROM
                (
                    SELECT
                        rcp_idwydarzenia,
                        rcp_czaswydarzenia,
                        rcp_typwydarzenia,
                        p_idpracownika,
                        lag(rcp_czaswydarzenia) OVER (PARTITION BY p_idpracownika ORDER BY rcp_czaswydarzenia DESC) AS czaswydarzenia_next
                    FROM
                        tb_rcp_wydarzenia
                    WHERE
                        (rcp_czaswydarzenia <= %s::timestamp + '22 hours 59 minutes 59 seconds'::interval)
                        AND (rcp_czaswydarzenia >= %s::timestamp + '00 hours 00 minutes 00 seconds'::interval)
                ) AS rcp
            WHERE
                (rcp.rcp_typwydarzenia IN (1, 2, 3, 4))
        ) AS pa_d
    LEFT OUTER JOIN tb_pracownicy AS wykpa_d ON (wykpa_d.p_idpracownika = pa_d.p_idpracownika)
    LEFT OUTER JOIN ts_dzialy AS dzwykpa_d ON (dzwykpa_d.dz_iddzialu = wykpa_d.dz_iddzialu)
    LEFT OUTER JOIN mvv.tb_pracownicy_mv AS m11wykpa_d ON (wykpa_d.p_idpracownika = m11wykpa_d.p_idpracownika)
    LEFT OUTER JOIN ts_stanowisko AS stwykpa_d ON (stwykpa_d.st_idstanowiska = wykpa_d.st_idstanowiska);
    """
    , (stockdate, stockdate, stockdate, stockdate,))


temp_data = cur.fetchall()

# Write the data to the "Data" sheet
for i in temp_data:
    page.append(i)

new_sheet = wb.active
new_sheet.title = 'Pracownicy'

# Use the same headers from the "Data" sheet for the "Pracownicy" sheet
new_sheet.append(pracownicy_headers)



# Execute a  query for the "Pracownicy" sheet
cur.execute("""
    SELECT DISTINCT wykpa_d.p_login,
        pa_d.dzien_pracy AS Data,
        wykpa_d.p_imie || ' ' || wykpa_d.p_nazwisko AS IiN,
        m11wykpa_d.p_value1527 AS BIBI,
        m11wykpa_d.p_value1447 AS department,
        m11wykpa_d.p_value1449 AS funkcja,
        m11wykpa_d.p_value1448 AS hcorag,
        m11wykpa_d.p_value1479 AS pozycja,
        wykpa_d.zm_idzmiany
    FROM
        (
            SELECT
                ((tpra.p_idpracownika::TEXT) || (tpra.pra_datastart::date)::TEXT) AS search_pole,
                tpra.p_idpracownika,
                SUM(tpra.pra_rbh) AS suma_rbh,
                SUM(tpra.pra_rbh * (tpra.pra_zaangazpracownika / 100)) AS suma_rbh_ef,
                MIN(tpra.pra_datastart) AS data_start,
                MAX(tpra.pra_datastop) AS data_stop,
                tpra.pra_datastart::DATE AS dzien_pracy,
                NULL::INT AS rcp_akcja,
                0 AS zrodlo
            FROM
                tg_praceall AS tpra
            WHERE
                (tpra.pra_flaga & 2 = 2)
                AND (tpra.pra_datastart::DATE <= %s)
                AND (tpra.pra_datastart::DATE >= %s)
            GROUP BY
                tpra.p_idpracownika,
                dzien_pracy
            UNION
            SELECT
                ((rcp.p_idpracownika::TEXT) || (rcp.rcp_czaswydarzenia::DATE)::TEXT || (rcp_idwydarzenia::TEXT)) AS search_pole,
                rcp.p_idpracownika,
                (CASE WHEN rcp.rcp_typwydarzenia IN (1, 3) THEN (EXTRACT(EPOCH FROM COALESCE(rcp.czaswydarzenia_next, now()) - rcp.rcp_czaswydarzenia) / 3600) ELSE 0 END) AS suma_rbh,
                0 AS suma_rbh_ef,
                rcp.rcp_czaswydarzenia AS data_start,
                COALESCE(rcp.czaswydarzenia_next, now()) AS data_stop,
                rcp.rcp_czaswydarzenia::DATE AS dzien_pracy,
                rcp.rcp_typwydarzenia AS rcp_akcja,
                1 AS zrodlo
            FROM
                (
                    SELECT
                        rcp_idwydarzenia,
                        rcp_czaswydarzenia,
                        rcp_typwydarzenia,
                        p_idpracownika,
                        lag(rcp_czaswydarzenia) OVER (PARTITION BY p_idpracownika ORDER BY rcp_czaswydarzenia DESC) AS czaswydarzenia_next
                    FROM
                        tb_rcp_wydarzenia
                    WHERE
                        (rcp_czaswydarzenia <= %s::timestamp + '22 hours 59 minutes 59 seconds'::interval)
                        AND (rcp_czaswydarzenia >= %s::timestamp + '00 hours 00 minutes 00 seconds'::interval)
                ) AS rcp
            WHERE
                (rcp.rcp_typwydarzenia IN (1, 2, 3, 4))
        ) AS pa_d
    LEFT OUTER JOIN tb_pracownicy AS wykpa_d ON (wykpa_d.p_idpracownika = pa_d.p_idpracownika)
    LEFT OUTER JOIN ts_dzialy AS dzwykpa_d ON (dzwykpa_d.dz_iddzialu = wykpa_d.dz_iddzialu)
    LEFT OUTER JOIN mvv.tb_pracownicy_mv AS m11wykpa_d ON (wykpa_d.p_idpracownika = m11wykpa_d.p_idpracownika)
    LEFT OUTER JOIN ts_stanowisko AS stwykpa_d ON (stwykpa_d.st_idstanowiska = wykpa_d.st_idstanowiska);
    """
    , (stockdate, stockdate, stockdate, stockdate,))


temp_pracownicy = cur.fetchall()

# Write the data to the "Pracownicy" sheet
for i in temp_pracownicy:
    new_sheet.append(i)


# Initialize dictionaries to store the lowest values for column J and the largest values for column K
lowest_values_j = {}
largest_values_k = {}

# Iterate through the data in the "Data" sheet
for row in page.iter_rows(min_row=2, values_only=True):  # Skip the header row
    login = row[2]  #  the "Login" column is at index 2 (0-based index)
    value_j = row[9]  #  column J is at index 9 (0-based index)
    value_k = row[10]  #  column K is at index 10 (0-based index)

    if login not in lowest_values_j or value_j < lowest_values_j[login]:
        lowest_values_j[login] = value_j

    if login not in largest_values_k or value_k > largest_values_k[login]:
        largest_values_k[login] = value_k

# Iterate through the "Pracownicy" sheet and update columns J and K
for row in new_sheet.iter_rows(min_row=2):
    login = row[0].value  #  the "Login" column is at index 0

    if login in lowest_values_j:
        row[9].value = lowest_values_j[login]  # Update column J with the lowest value

    if login in largest_values_k:
        row[10].value = largest_values_k[login]  # Update column K with the largest value
        
# Calculate the "Czas Pracy" and set the values in column L
for row in new_sheet.iter_rows(min_row=2):
    start_pracy = row[9].value  #  column J is at index 9
    koniec_pracy = row[10].value  #  column K is at index 10

    if start_pracy and koniec_pracy:
        start_time = datetime.strptime(start_pracy, '%H:%M')
        end_time = datetime.strptime(koniec_pracy, '%H:%M')

        if start_time > end_time:
            # Handle cases where start time is greater than end time (spans to the next day)
            end_time += timedelta(days=1)

        czas_pracy = end_time - start_time

        # Format the result as hours:minutes
        czas_pracy_str = f'{czas_pracy.days * 24 + czas_pracy.seconds // 3600:02}:{(czas_pracy.seconds // 60) % 60:02}'

        row[11].value = czas_pracy_str
        
# Generate values in column M of the "Pracownicy" sheet
wsad_column_M_values = {}
for row in page.iter_rows(min_row=2):
    login = row[2].value  #  "Login" is in column C (0-based index)

    if row[12].value == 3:  #  column M is at index 12
        if login not in wsad_column_M_values:
            wsad_column_M_values[login] = row[11].value  # Copy the value from column L

for row in new_sheet.iter_rows(min_row=2):
    login = row[0].value  #  "Login" is in column A (0-based index)

    if login in wsad_column_M_values:
        row[12].value = wsad_column_M_values[login]  # Set the value in column M
        
###

for row in new_sheet.iter_rows(min_row=2):
    start_pracy = row[9].value  #  column J is at index 9

    if start_pracy:
        start_time = datetime.strptime(start_pracy, '%H:%M')

        if start_time <= datetime.strptime('06:00', '%H:%M'):
            row[14].value = 'OK'  #  you want to add the formula to column O (index 14)
        else:
            row[14].value = start_pracy  # Copy the value from column J
            

# Iterate through the "Pracownicy" sheet and add the formula to column P
for row in new_sheet.iter_rows(min_row=2):
    czas_pracy = row[11].value  #  column L is at index 11
    koniec_pracy = row[10].value  #  column K is at index 10

    if czas_pracy:
        end_time = datetime.strptime(czas_pracy, '%H:%M')

        if end_time < datetime.strptime('8:00', '%H:%M'):
            row[15].value = koniec_pracy  # Copy the value from column K
        else:
            row[15].value = 'OK'  # Otherwise, paste "OK" to column P
   
         

# Iterate through the "Pracownicy" sheet and add the formula to column Q
for row in new_sheet.iter_rows(min_row=2):
    czas_pracy = row[11].value  # column L is at index 11

    if czas_pracy:
        hours, minutes = map(int, czas_pracy.split(':'))
        total_minutes = hours * 60 + minutes

        if total_minutes < 8 * 60:  # Check if the total time is less than 8 hours
            row[16].value = 'Poniżej 8h'  # If less than 8 hours, paste "poniżej 8h" to column Q
        else:
            row[16].value = 'OK'  # Otherwise, paste "OK" to column Q

# Iterate through the "Pracownicy" sheet and add the formula to column R
for row in new_sheet.iter_rows(min_row=2):
    value_M = row[12].value  #  column M is at index 12

    if value_M:
        row[17].value = 'OK'  # If there is a value in column M, paste "OK" in column R
    else:
        row[17].value = 'brak przerwy'  # If there's no value in column M, paste "brak przerwy" in column R
        
# Iterate through the "Pracownicy" sheet and add the emojis to column S
for row in new_sheet.iter_rows(min_row=2):
    value_O = row[14].value  #  column O is at index 14
    value_P = row[15].value  #  column P is at index 15
    value_Q = row[16].value  #  column Q is at index 16
    value_R = row[17].value  #  column R is at index 17
    
    cell_S = row[18]  #  column S is at index 18

    if all(value in ['OK'] for value in [value_O, value_Q, value_R, value_P]):
        # If all three columns have "OK", add a checkmark (✓) emoji to column S
        cell_S.font = Font(color="008000")  # Green in RGB
        row[18].value = "✔️"
        

    else:
        # If any column does not have "OK," add an X mark (✗) emoji to column S
        cell_S.font = Font(color="FF0000")  # Red in RGB
        row[18].value = "❌"

###

# Iterate through the rows to add borders
for row in new_sheet.iter_rows():
    for cell in row:
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


# Define a background color fill for the header row
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Define a font style for making the header row bolder
header_font = Font(bold=True)

# Apply the background color fill to the header row
for row in new_sheet.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        
# Iterate through the columns to adjust column widths
for column in new_sheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)  # Get the column letter

    # Find the maximum length of the data in the column
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass

    adjusted_width = (max_length + 2)  # Add some padding
    new_sheet.column_dimensions[column_letter].width = adjusted_width

###



# Save the workbook with the new sheet names
workbook_name = 'Raport.xlsx'
wb.save(workbook_name)
wb.close()

print(workbook_name)
cur.close()
conn.close()
